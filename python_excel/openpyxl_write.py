import openpyxl
import datetime
from openpyxl.styles import Font, colors, Alignment
#实例化
workbook = openpyxl.Workbook()
# 激活 worksheet
sheet=workbook.active
#写入数据
sheet['A1']='python'
sheet['B1']='javascript'
#写入时间
sheet['A2'] = datetime.datetime.now().strftime("%Y-%m-%d")
# 第2行行高
sheet.row_dimensions[2].height = 40
# B列列宽
sheet.column_dimensions['B'].width = 30
# 设置A1中的数据垂直居中和水平居中
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
# 下面的代码指定了等线24号，加粗斜体，字体颜色黄色。直接使用cell的font属性，将Font对象赋值给它。
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color='00FFBB00', bold=True)
sheet['B1'].font = bold_itatic_24_font
# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('A2:B2') # 合并一行中的几个单元格
# 拆分单元格
# sheet.unmerge_cells('A2:B2')
#保存
workbook.save('new.xlsx')
